import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter import Radiobutton
import openpyxl, pyinputplus as pyip, datetime
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import warnings
import time

def fileSearch(event):
    global greeting, fileName, enterFile, fileDesc, showFileName, yesButton, noButton
    fileName = askopenfilename()
    enterFile.pack_forget()
    fileButton.pack_forget()
    greeting.pack_forget()
    
    fileDesc = tk.Label(text="Would you like to use file: ")
    showFileName = tk.Label(text=fileName)
    yesButton = tk.Button(text="Yes")
    yesButton.bind("<Button-1>",openFile)
    noButton = tk.Button(text="No")
    noButton.bind("<Button-1>",clearWindow)
    fileDesc.pack()
    showFileName.pack()
    yesButton.pack()
    noButton.pack()
    
    return fileName

def clearWindow(event):
    global fileName, fileDesc, showFileName, yesButton, noButton
    fileDesc.pack_forget()
    showFileName.pack_forget()
    yesButton.pack_forget()
    noButton.pack_forget()
    fileSearch(event)

def pressOk(event):
    global wrongFileType, okButton
    wrongFileType.pack_forget()
    okButton.pack_forget()
    fileSearch(event)

def openFile(event):
    global wb, sheet, frame, indivYears, validProjectStatus, radio, fileName, fileDesc, showFileName, yesButton, noButton, wrongFileType, okButton, fiscalYearLabel, options, indivYears, okButton
    fileDesc.pack_forget()
    showFileName.pack_forget()
    yesButton.pack_forget()
    noButton.pack_forget()
    wrongFileType = tk.Label(text="Incorrect File Type. Try Again")

    fileExtension = fileName[-4:]
    print(fileExtension)
    if (fileExtension == "xlsx"):
        openingLabel = tk.Label(text="Opening your file...")
        openingLabel.pack()
        time.sleep(3)
    
    else:
        wrongFileType.pack()
        okButton = tk.Button(text="Ok")
        okButton.bind("<Button-1>",pressOk)
        okButton.pack()

    warnings.simplefilter(action='ignore', category=UserWarning)
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    validProjectStatus = ["Completed", "In Progress", "Waiting for sign-off", "Duplicate - Additional Partner"]
    currentDate = datetime.datetime.now()
    currentYear = currentDate.year

    
    #Gathers all fiscal years present in spreadsheet
    allFiscalYears = []
    for row in range(5,sheet.max_row):
        rowValue = sheet["BA" + str(row)].value
        if (str(rowValue) not in allFiscalYears):
            allFiscalYears.append(rowValue)

    indivYears = []     
    for year in allFiscalYears:
        if year in indivYears or year == None:
            continue
        else:
            indivYears.append(year)
    indivYears.sort()

    
    #Asks user to input Fiscal Year and whether its interim or year-end
    openingLabel.pack_forget()

    frame = Frame(window)
    frame.pack()
    
    fiscalYearLabel = tk.Label(frame, text="Please select a fiscal year")
    fiscalYearLabel.pack()
    radio = IntVar()
    
    for i in range(0,len(indivYears)):
        options = Radiobutton(frame, text=indivYears[i], variable=radio, value = i)
        options.pack()

    
    okButton = tk.Button(frame, text="Ok")
    okButton.bind("<Button-1>", getReportType)
    okButton.pack()

def getReportType(event):
    global radio, frame, fiscalYearChoice, fiscalYearLabel, options, indivYears, okButton, typeOfReport, interim, yearEnd

    fiscalYearLabel.pack_forget()
    options.pack_forget()
    okButton.pack_forget()

    frame.destroy()
    typeOfReport = tk.Label(text="Which scorecard is being created?")
    typeOfReport.pack()
    interim = Radiobutton(text="Interim", variable=radio, value = "0")
    interim.pack()
    yearEnd = Radiobutton(text="Year-End", variable=radio, value = "1")
    yearEnd.pack()
    fiscalYearChoice = radio.get()
    okButton = tk.Button(text="Ok")
    okButton.bind("<Button-1>", gatherData)
    okButton.pack()

def gatherData(event):
    global validProjectStatus, okButton, typeOfReport, interim, yearEnd, wb, sheet, fiscalYearChoice, indivYears
    typeOfReport.pack_forget()
    interim.pack_forget()
    yearEnd.pack_forget()
    okButton.pack_forget()
    timeOfYearChoice = radio.get()

    fiscalYear = indivYears[fiscalYearChoice]

    timeOfYearOptions = ["Interim","Year-End"]
    timeOfYear = timeOfYearOptions[timeOfYearChoice]
    
    gatheringLabel = tk.Label(text="Gathering Data...")

    facultyColumns = ["X","Y","Z","AA","AB"]
    studentColumns = ["AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX"]          
    currentFiscalYearRows = []

    #Grabs rows for everything in the selected fiscal year
    for row in range(5,sheet.max_row):
        rowValue = sheet["BA" + str(row)].value
        if (str(rowValue) == str(fiscalYear)):
            currentFiscalYearRows.append(row)

    #Based off the project status, grab only the relevant rows
    relevantRows = []
    for row in currentFiscalYearRows:
        rowProjStat = sheet["AZ" + str(row)].value
        if (rowProjStat in validProjectStatus):
            relevantRows.append(row)

    projects = []
    companies = []
    faculty = []
    students = []

    #Creates lists of all project numbers and company names within the relevant rows
    for row in relevantRows:
        projectNum = sheet["A" + str(row)].value
        projects.append(projectNum)
        companyName = sheet["C" + str(row)].value
        companies.append(companyName)

    #Creates list of all faculty names within relevant rows
    for column in facultyColumns:
        for row in relevantRows:
            facultyName = sheet[column + str(row)].value
            faculty.append(facultyName)

    #Creates list of all student names within relevant rows
    for column in studentColumns:
        for row in relevantRows:
            studentName = sheet[column + str(row)].value
            students.append(studentName)

    indivProjects = []
    indivCompanies = []
    indivFaculty = []
    indivStudents = []

    #Gets rid of all None values and duplicates
    for number in projects:
        if number in indivProjects or number == None:
            continue
        else:
            indivProjects.append(number)

    for name in companies:
        if name in indivCompanies or name == None:
            continue
        else:
            indivCompanies.append(name)

    for name in faculty:
        if name in indivFaculty or name == None:
            continue
        else:
            indivFaculty.append(name)

    for name in students:
        if name in indivStudents or name == None:
            continue
        else:
            indivStudents.append(name)

    wb.close()

    #Create new Scorecard
    newScoreCard = openpyxl.Workbook()
    sheet = newScoreCard.active
    rowTitles = ["Metrics","Projects","Faculty Researchers","Student Researchers","Industry Partners"]
    data = [len(indivProjects), len(indivFaculty), len(indivStudents), len(indivCompanies)]

    #Add Row Titles and bold
    for row in range(0,len(rowTitles)):
        sheet["B" + str(row + 2)] = rowTitles[row]
        sheet["B" + str(row + 2)].font = Font(bold=True)

    #Add data Column info
    sheet["C2"] = str(fiscalYear) + " (" + str(timeOfYear) + ")"
    sheet["C2"].font = Font(bold=True)
    for row in range(0,len(data)):
        sheet["C" + str(row + 3)] = data[row]

    #Style the sheet
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    greyFill = PatternFill(fill_type="solid", start_color='808080', end_color='808080')
    lightGreyFill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    blueFill = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")
    lightBlueFill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    sheet["B2"].fill = greyFill
    sheet["B2"].font = Font(bold=True, color='FFFFFF')
    sheet["C2"].font = Font(bold=True, color='FFFFFF')

    for row in range(3,7):
        sheet["B" + str(row)].fill = lightGreyFill

    if (timeOfYear == "Interim"):
        sheet["C2"].fill = blueFill
        for row in range(3,7):
            sheet["C" + str(row)].fill = lightBlueFill
    else:
        sheet["C2"].fill = greyFill
        
    #save Excel file

    if (os.path.exists(r"C:\ScoreCards") == False):
        os.makedirs(r"C:\ScoreCards")
    newScoreCard.save("C:\\ScoreCards\\BoG Scorecard " + fiscalYear[0] + fiscalYear[1] + fiscalYear[2] + fiscalYear [3] + "-" + fiscalYear[5] + fiscalYear[6] + " " + str(timeOfYear) + ".xlsx")

    gatheringLabel.pack_forget()

    processCompletedLabel = tk.Label(text="Process Completed! You may now close this window.")
    processCompletedLabel.pack()
    print("Process Completed!")
    Button(window, text="Quit", command=window.destroy).pack()

        

window = tk.Tk()
window.geometry("500x200")
greeting = tk.Label(text="Welcome")
enterFile = tk.Label( text="Please select the master spreadsheet")
fileButton = tk.Button(text="Search")
fileButton.bind("<Button-1>",fileSearch)

greeting.pack()
enterFile.pack()
fileButton.pack()



window.mainloop()
